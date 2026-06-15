"""Funciones reutilizables para exportar reportes a PDF y Excel."""
from django.http import HttpResponse
from django.utils import timezone


# ─────────────────────────────────────────
# PDF con reportlab
# ─────────────────────────────────────────

def generar_pdf_tabla(titulo, subtitulo, encabezados, filas, nombre_archivo):
    """
    Genera un PDF con una tabla de datos.
    - titulo: str
    - subtitulo: str (rango de fechas u otro contexto)
    - encabezados: list[str]
    - filas: list[list] — cada fila es una lista de valores
    - nombre_archivo: str sin extensión
    """
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        'titulo', parent=styles['Title'], fontSize=16, spaceAfter=6
    )
    estilo_subtitulo = ParagraphStyle(
        'subtitulo', parent=styles['Normal'], fontSize=10,
        textColor=colors.grey, spaceAfter=12
    )
    estilo_pie = ParagraphStyle(
        'pie', parent=styles['Normal'], fontSize=8,
        textColor=colors.grey, alignment=1
    )

    elementos = []

    # Título y subtítulo
    elementos.append(Paragraph(titulo, estilo_titulo))
    elementos.append(Paragraph(subtitulo, estilo_subtitulo))
    elementos.append(Spacer(1, 0.3 * cm))

    # Tabla
    data = [encabezados] + [list(map(str, fila)) for fila in filas]
    tabla = Table(data, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#212529')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elementos.append(tabla)

    # Pie de página
    elementos.append(Spacer(1, 0.5 * cm))
    ahora = timezone.localtime(timezone.now())
    elementos.append(Paragraph(
        f'Generado el {ahora:%d/%m/%Y %H:%M}',
        estilo_pie
    ))

    doc.build(elementos)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    ahora_str = ahora.strftime('%Y%m%d_%H%M')
    response['Content-Disposition'] = (
        f'attachment; filename="{nombre_archivo}_{ahora_str}.pdf"'
    )
    return response


# ─────────────────────────────────────────
# Excel con openpyxl
# ─────────────────────────────────────────

def generar_excel_tabla(titulo, encabezados, filas, nombre_archivo, columnas_moneda=None):
    """
    Genera un Excel con una tabla de datos.
    - columnas_moneda: list[int] — índices de columnas con formato de moneda (base 0)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import io

    columnas_moneda = columnas_moneda or []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo[:31]  # Excel limita a 31 caracteres

    # Encabezados
    fill_header = PatternFill(start_color='212529', end_color='212529', fill_type='solid')
    font_header = Font(color='FFFFFF', bold=True)

    for col_idx, encabezado in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col_idx, value=encabezado)
        celda.fill = fill_header
        celda.font = font_header
        celda.alignment = Alignment(horizontal='center')

    # Filas de datos
    for row_idx, fila in enumerate(filas, start=2):
        for col_idx, valor in enumerate(fila, start=1):
            celda = ws.cell(row=row_idx, column=col_idx, value=valor)
            if (col_idx - 1) in columnas_moneda:
                celda.number_format = '"$"#,##0.00'

    # Ajustar ancho de columnas
    for col_idx, _ in enumerate(encabezados, start=1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for celda in row:
                try:
                    if celda.value:
                        max_len = max(max_len, len(str(celda.value)))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    ahora = timezone.localtime(timezone.now())
    ahora_str = ahora.strftime('%Y%m%d_%H%M')

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="{nombre_archivo}_{ahora_str}.xlsx"'
    )
    return response